import openpyxl
from backend.app.db.session import SessionLocal, engine
from backend.app.db.models import (
    Base, User, Category, Vendor, PurchaseRequisition, PurchaseOrder, GRN, Claim,
    SupplierFeedback, PRStatus, UserRole, MarketIntelligence, VendorDiscovery, SavingsTracker,
    MaterialServiceMaster, CategoryWorkbook, CategoryWorkbookSection, MasterContractClause
)
from datetime import datetime

def parse_id(val):
    if not val: return None
    try:
        # Expected format like CAT_001 or VEND_005 or just 5
        s_val = str(val)
        if '_' in s_val:
            return int(s_val.split('_')[-1])
        return int(s_val)
    except:
        return None

def get_row_val(row, idx, default=None):
    if not row or idx >= len(row):
        return default
    return row[idx]

def migrate_full():
    Base.metadata.drop_all(bind=engine) # Start fresh for clean migration
    Base.metadata.create_all(bind=engine)
    
    path = r"c:\Users\Moiz\Desktop\Procurement\backend\Database\Database for Procurement.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    db = SessionLocal()
    
    try:
        # 1. Categories
        print("Migrating Categories...")
        if 'Category Master' in wb.sheetnames:
            sheet = wb['Category Master']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                db.merge(Category(
                    id=parse_id(row[0]), 
                    name=get_row_val(row, 1, ""), 
                    description=get_row_val(row, 2, "")
                ))
            db.commit()

        # 2. Users (Employee Master)
        print("Migrating Users...")
        if 'Employee Master' in wb.sheetnames:
            sheet = wb['Employee Master']
            role_map = {
                'Chief Procurement Officer': UserRole.CPO,
                'Procurement Category Manager': UserRole.CATEGORY_MANAGER,
                'Sourcing Analyst': UserRole.ANALYST,
                'Requisitioner': UserRole.REQUESTER,
                'Supplier Relationship Manager': UserRole.SRM
            }
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                db.merge(User(
                    id=parse_id(row[0]), 
                    name=get_row_val(row, 1, ""), 
                    email=get_row_val(row, 3, ""), 
                    role=role_map.get(get_row_val(row, 5), UserRole.ANALYST)
                ))
            db.commit()

        # 3. Vendors (Vendor Master)
        print("Migrating Vendors...")
        if 'Vendor Master' in wb.sheetnames:
            sheet = wb['Vendor Master']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                cat_name = get_row_val(row, 3)
                cat = db.query(Category).filter(Category.name == cat_name).first() if cat_name else None
                risk_score_raw = get_row_val(row, 19) # Overall_Risk_Score is at 30 in some sheets, but Rating is at 19 in headers_log
                db.merge(Vendor(
                    id=parse_id(row[0]), 
                    name=get_row_val(row, 1, ""), 
                    category_id=cat.id if cat else None,
                    risk_score=(float(risk_score_raw) if risk_score_raw else 5)/10.0, 
                    esg_score=0.75
                ))
            db.commit()

        # 4. PRs (Purchase Requisition)
        print("Migrating PRs...")
        if 'Purchase Requisition' in wb.sheetnames:
            sheet = wb['Purchase Requisition']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                
                raw_status = str(get_row_val(row, 11) or "CREATED").strip().upper()
                status_enum = PRStatus.APPROVED
                if raw_status in ["CREATED", "APPROVED", "SOURCING", "PO_CREATED", "REJECTED"]:
                    status_enum = PRStatus(raw_status)
                elif raw_status == "CLOSED":
                    status_enum = PRStatus.PO_CREATED

                amount_raw = str(get_row_val(row, 9) or 0).replace(',', '')
                try: amount = float(amount_raw)
                except ValueError: amount = 0.0

                created_at = get_row_val(row, 1)
                approved_at = get_row_val(row, 12)
                required_by = get_row_val(row, 13)

                db.merge(PurchaseRequisition(
                    id=parse_id(row[0]), 
                    pr_number=str(row[0]).strip(),
                    requester_id=parse_id(get_row_val(row, 2)), 
                    department=str(get_row_val(row, 5) or ""),
                    category_id=parse_id(get_row_val(row, 6)),
                    ms_id=str(get_row_val(row, 7) or ""),
                    description=str(get_row_val(row, 8) or ""),
                    amount=amount, 
                    currency=str(get_row_val(row, 10) or "INR"),
                    status=status_enum, 
                    created_at=created_at if isinstance(created_at, datetime) else datetime.utcnow(),
                    approved_at=approved_at if isinstance(approved_at, datetime) else None,
                    required_by=required_by if isinstance(required_by, datetime) else None,
                    location_id=str(get_row_val(row, 14) or ""),
                    bu_id=str(get_row_val(row, 15) or ""),
                    priority=str(get_row_val(row, 16) or "Medium"),
                    current_owner="System", 
                    sla_days=10
                ))
            db.commit()

        # 5. POs (Purchase Order)
        print("Migrating POs...")
        if 'Purchase Order' in wb.sheetnames:
            sheet = wb['Purchase Order']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                # Indices from headers_log.txt: [12] PO_Amount_INR, [13] Currency, [16] Status
                amount_raw = get_row_val(row, 12, 0)
                try: amount = float(amount_raw) if amount_raw else 0.0
                except: amount = 0.0
                
                db.merge(PurchaseOrder(
                    id=parse_id(row[0]), 
                    po_number=str(row[0]), 
                    pr_id=parse_id(get_row_val(row, 2)),
                    vendor_id=parse_id(get_row_val(row, 4)), 
                    amount=amount, 
                    currency=get_row_val(row, 13, "INR"),
                    status=get_row_val(row, 16, "Issued"),
                    mode_of_purchase=get_row_val(row, 11, "Direct")
                ))
            db.commit()

        # 6. GRNs
        print("Migrating GRNs...")
        if 'GRN' in wb.sheetnames:
            sheet = wb['GRN']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                # Indices from headers_log.txt: [10] Received_Date, [8] Received_Quantity, [15] Quality_Status
                received_date = get_row_val(row, 10)
                db.merge(GRN(
                    id=parse_id(row[0]), 
                    grn_number=str(row[0]), 
                    po_id=parse_id(get_row_val(row, 2)),
                    received_date=received_date if isinstance(received_date, datetime) else datetime.utcnow(),
                    received_amount=get_row_val(row, 8, 0), 
                    quality_status=get_row_val(row, 15, "Accepted")
                ))
            db.commit()

        # 7. Claims
        print("Migrating Claims...")
        if 'Claims' in wb.sheetnames:
            sheet = wb['Claims']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                # Indices from headers_log.txt: [8] Claim_Amount_INR, [7] Claim_Type, [11] Claim_Status
                db.merge(Claim(
                    id=parse_id(row[0]), 
                    claim_number=str(row[0]), 
                    po_id=parse_id(get_row_val(row, 1)),
                    vendor_id=parse_id(get_row_val(row, 2)), 
                    amount=get_row_val(row, 8, 0), 
                    type=get_row_val(row, 7, "Quality"),
                    status=get_row_val(row, 11, "Open")
                ))
            db.commit()

        # 8. Vendor Ratings (Feedback)
        print("Migrating Vendor Ratings (Feedback)...")
        if 'Vendor Rating' in wb.sheetnames:
            sheet = wb['Vendor Rating']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                # Indices from headers_log.txt: [18] Overall_Vendor_Score, [23] Feedback_Comments
                db.add(SupplierFeedback(
                    vendor_id=parse_id(get_row_val(row, 2)), 
                    rating=get_row_val(row, 18, 3.0),
                    comment=get_row_val(row, 23, "No comment"), 
                    review_date=datetime.utcnow()
                ))
            db.commit()

        # 9. Intelligence (Market Intel)
        print("Migrating Market Intelligence...")
        if 'Market Intelligence' in wb.sheetnames:
            sheet = wb['Market Intelligence']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                # Indices from headers_log.txt: [6] Impact_Category_Name, [1] Information_Title, [11] Impact_Level
                db.add(MarketIntelligence(
                    category_name=get_row_val(row, 6, "General"), 
                    commodity=get_row_val(row, 1, "Unknown"), 
                    region=get_row_val(row, 8, "Global"),
                    current_price="1200", 
                    price_unit="USD", 
                    trend=get_row_val(row, 11, "Stable"), 
                    forecast_6m="N/A"
                ))
            db.commit()

        # 10. Vendor Discovery
        print("Migrating Vendor Discovery...")
        if 'Vendor Discovery' in wb.sheetnames:
            sheet = wb['Vendor Discovery']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                # Indices from headers_log.txt: [2] Vendor_Name, [4] Category_Name, [7] Description
                db.add(VendorDiscovery(
                    vendor_name=get_row_val(row, 2, "New Vendor"), 
                    category_name=get_row_val(row, 4, "General"), 
                    capabilities=get_row_val(row, 7, "Unknown"),
                    geographical_presence=get_row_val(row, 8, "Global"), 
                    scouted_by="System"
                ))
            db.commit()

        # 11. Category Workbooks
        print("Migrating Category Workbooks...")
        if 'Category Workbook' in wb.sheetnames:
            sheet = wb['Category Workbook']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                cat_id_val = get_row_val(row, 1)
                cat = db.query(Category).filter(Category.id == parse_id(cat_id_val)).first() if cat_id_val else None
                db.add(CategoryWorkbook(
                    workbook_id=str(row[0]),
                    category_id=cat.id if cat else None,
                    description=get_row_val(row, 3, ""),
                    owner=get_row_val(row, 4, ""),
                    status=get_row_val(row, 6, "")
                ))
            db.commit()

        # 12. Category Workbook Sections
        print("Migrating Workbook Sections...")
        if 'Category Workbook Sections' in wb.sheetnames:
            sheet = wb['Category Workbook Sections']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                section_num_raw = get_row_val(row, 4)
                db.add(CategoryWorkbookSection(
                    section_id=str(row[0]),
                    workbook_id=str(get_row_val(row, 1)),
                    section_number=int(section_num_raw) if section_num_raw else 0,
                    section_name=str(get_row_val(row, 5, "")),
                    content=str(get_row_val(row, 6, ""))
                ))
            db.commit()

        # 13. Material Service Master
        print("Migrating Material Master...")
        if 'Material Service Master' in wb.sheetnames:
            sheet = wb['Material Service Master']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                cat_id_val = get_row_val(row, 3)
                cat = db.query(Category).filter(Category.id == parse_id(cat_id_val)).first() if cat_id_val else None
                
                # Parse numeric values safely
                price = 0
                try:
                    price_raw = get_row_val(row, 10)
                    price = float(price_raw) if price_raw else 0
                except:
                    pass
                
                annual_qty = 0
                try:
                    annual_qty_raw = get_row_val(row, 19)
                    annual_qty = float(annual_qty_raw) if annual_qty_raw else 0
                except:
                    pass

                db.add(MaterialServiceMaster(
                    ms_id=str(row[0]),
                    type=get_row_val(row, 1, ""),
                    sub_type=get_row_val(row, 2, ""),
                    category_id=cat.id if cat else None,
                    material_name=get_row_val(row, 7, "Unknown"),
                    description=get_row_val(row, 7, ""),
                    grade=get_row_val(row, 8, ""),
                    price=price,
                    uom=get_row_val(row, 15, ""),
                    annual_quantity=annual_qty
                ))
            db.commit()

        # 14. Master Contract Clauses
        print("Migrating Master Contract Clauses...")
        if 'Master Contract Clause' in wb.sheetnames:
            sheet = wb['Master Contract Clause']
            counter = 1
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                db.add(MasterContractClause(
                    clause_id=str(row[0]),              
                    clause_type=str(get_row_val(row, 3) or ''),      
                    clause_name=str(get_row_val(row, 1) or ''),      
                    description=str(get_row_val(row, 1) or ''),      
                    standard_language=str(get_row_val(row, 2) or ''),
                    applicability=str(get_row_val(row, 4, "") or ""),
                    risk_level=str(get_row_val(row, 5, "") or ""),
                    is_mandatory=str(get_row_val(row, 6, "") or ""),
                ))
                counter += 1
            db.commit()
            print(f"  -> Inserted {counter-1} MCC clauses.")

        print("Full production migration completed successfully!")

    except Exception as e:
        import traceback
        print(f"Migration Error: {e}")
        print(traceback.format_exc())
        db.rollback()
    finally:
        wb.close()
        db.close()

if __name__ == "__main__":
    migrate_full()
